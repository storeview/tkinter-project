from openpyxl import load_workbook
from openpyxl import Workbook
from copy import copy

datas = []
imported_workbooks = []
selected_columns = []
column_titles = []
columns_name = []


def add_file(filenames):
    for filename in filenames:
        if filename in imported_workbooks:
            print("filename ALREADY selected")
        else:
            ret = analysis_excel(filename, get_file_type(filename))
            datas.append({"filename": ret[0], "sheetnames": ret[1], "selected": ret[2], 
                              "rows": ret[3], "columns": ret[4]})
            imported_workbooks.append(filename)
    return datas


def delete_file(index):
    print(f"removed: {datas[index]}")
    imported_workbooks.remove(datas[index]["filename"])
    datas.remove(datas[index])
    return datas


def get_columns():
    first_one = datas[0]
    return   get_columns_by_filetype(first_one, get_file_type(first_one["filename"]))


def export_file(filename, selected_column, seperate_conditions):
    return export_file_by_filetype(filename, selected_column, seperate_conditions, get_file_type(filename))


def get_file_type(filename):
    if len(filename.split(".xlsx")) > 1:
        return "xlsx"
    elif len(filename.split(".xls")) > 1:
         return "xls"
    return ""


def analysis_excel(filename, filetype):
    wb = None
    sh = None
    if filetype == "xls":
        print("xls")
    elif filetype == "xlsx":
        wb = load_workbook(filename=filename)
        sh = wb.get_sheet_by_name(wb.sheetnames[0])
    return [filename, wb.sheetnames, 0, sh.max_row, sh.max_column]


def get_columns_by_filetype(first_one, filetype):
    global column_titles
    global columns_name

    if filetype == "xlsx":
        selected_sheet_index = first_one["selected"]
        max_column = first_one["columns"]
        wb = load_workbook(filename=first_one["filename"])
        sh = wb.get_sheet_by_name(wb.sheetnames[selected_sheet_index])
        column_titles = [sh.cell(1, i).value for i in range(1, max_column+1)]
        columns_name = [sh.cell(2, i).value for i in range(1, max_column+1)]
        #+1的是代表【全选】
        selected_columns = [1 for i in range(len(column_titles)+1)]

        return [column_titles, selected_columns]
    else:
        return []


def get_column_name_by_index(index):
    column_names = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z"]
    return column_names[index-1]

def export_file_by_filetype(filename, selected_column, seperate_conditions, filetype):
    #filename：导出文件的名称
    #selected_column：选择导出的列
    #seperate_conditions：条件列中的过滤项
    if filetype == "xlsx":
        new_excel = Workbook()
        sheet = new_excel.active
        sheet.title = "Sheet1"

        first_one = datas[0]
        old_excel = load_workbook(filename=first_one["filename"])
        selected_sheet_index = first_one["selected"]
        old_sheet = old_excel.get_sheet_by_name(old_excel.sheetnames[selected_sheet_index])
        #复制列宽和行高
        
        
        #列计数，保证选中的列从最左边开始排
        column_num = 1
        for i in range (len(selected_column)-1):
            if selected_column[i] == 1:
                copy_cell_value_and_all_style(old_sheet.cell(1, i+1), sheet.cell(1, column_num))
                copy_cell_value_and_all_style(old_sheet.cell(2, i+1), sheet.cell(2, column_num))
                sheet.column_dimensions[get_column_name_by_index(column_num)].width = old_sheet.column_dimensions[get_column_name_by_index(i+1)].width
                column_num +=  1
        #从第三行开始输出数据
        index = 3
        for excel in datas:
            wb = load_workbook(filename=excel["filename"])
            sh = wb.get_sheet_by_name(wb.sheetnames[excel["selected"]])
            rows = list(sh.rows)
            for row in rows[2:]:
                #列计数，保证选中的列从最左边开始排
                column_num = 1
                for i in range (len(selected_column)-1):
                    if selected_column[i] == 1:
                        copy_cell_value_and_all_style(row[i], sheet.cell(index, column_num))
                        column_num += 1
                index += 1
        
        sheet.row_dimensions[1].height = old_sheet.row_dimensions[1].height

        new_excel.save(filename)
        return "导出成功！" 
    
    elif filetype == "xls":
        return "导出成功！"

def copy_cell_value_and_all_style(target, copy_cell):
    copy_cell.value = target.value
    copy_cell.font = copy(target.font)
    copy_cell.fill = copy(target.fill)
    copy_cell.alignment = copy(target.alignment)
    copy_cell.border = copy(target.border)
    copy_cell.number_format = copy(target.number_format)
    copy_cell.hyperlink = copy(target.hyperlink)

if __name__ == "__main__":
    filenames = ('D:/Windows数据移动到此文件夹/Desktop/测试用例导出.xlsx', 'D:/Windows数据移动到此文件夹/Desktop/测试用例导出 - 副本.xlsx')
    add_file(filenames)
    get_columns()